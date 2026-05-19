from openpyxl import load_workbook

class ExcelReader:
    """Custom library to read Excel files using openpyxl"""
    
    def read_excel_data(self, file_path):
        """
        Read data from Excel file and return as list of dictionaries.
        Assumes first row contains headers.
        """
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # Get headers from the first row
        headers = [cell.value for cell in sheet[1]]
        
        # Read data rows
        data = []
        # min_row=2 starts reading from the second row to skip headers
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for idx, value in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = value
            data.append(row_data)
            
        workbook.close()
        return data